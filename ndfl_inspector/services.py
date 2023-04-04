from collections import namedtuple
from typing import List, Optional

import openpyxl
from openpyxl.cell.cell import Cell
from openpyxl.styles import PatternFill
from openpyxl.worksheet.worksheet import Worksheet


TAX_THRESHOLD = 5000000
PERCENT_BEFORE_THRESHOLD = 0.13
PERCENT_AFTER_THRESHOLD = 0.15
RIGHT_COLOR = '00FF00'
WRONG_COLOR = 'FF0000'
DEVIATION_COLUMN_NUM = 6


def read_income_excel(filepath: str,
                      start_row: int,
                      column_nums: dict[str, int]) -> List[namedtuple]:
    income_file = openpyxl.load_workbook(filepath, read_only=True)
    income_sheet = income_file.active
    columns = ('branch', 'employee', 'ndfl_base', 'custom_total')
    EmployeeNDFL = namedtuple('EmployeeNDFL', columns)

    income_ndfl_table = []
    for row in income_sheet.iter_rows(values_only=True, min_row=start_row):
        ndfl_row = EmployeeNDFL(
            row[column_nums['branch_column_num'] - 1],
            row[column_nums['employee_column_num'] - 1],
            row[column_nums['tax_base_column_num'] - 1],
            row[column_nums['custom_total_column_num'] - 1],
        )
        if ndfl_row.employee:
            income_ndfl_table.append(ndfl_row)

    income_file.close()
    return income_ndfl_table


def generate_outcome_excel(income_ndfl_table: List[namedtuple]) -> openpyxl.Workbook:
    outcome_workbook = openpyxl.Workbook()
    outcome_sheet = outcome_workbook.active
    for row_number, row in enumerate(income_ndfl_table, 1):
        calculation = _calculate_ndfl(row.ndfl_base)

        deviation = _calculate_deviation(row.custom_total, calculation)

        result_row = (
            row.branch,
            row.employee,
            row.ndfl_base,
            row.custom_total,
            calculation,
            deviation,
        )
        outcome_sheet.append(result_row)

        deviation_cell = outcome_sheet.cell(
            row=row_number,
            column=DEVIATION_COLUMN_NUM,
        )
        _colorize_deviation_cell(deviation_cell, deviation)


    _generate_outcome_header(outcome_sheet)

    return outcome_workbook


def _calculate_ndfl(ndfl_base: Optional[float] = None) -> int:
    if not ndfl_base:
        return 0
    if ndfl_base < TAX_THRESHOLD:
        return round(ndfl_base * PERCENT_BEFORE_THRESHOLD)
    return round(ndfl_base * PERCENT_AFTER_THRESHOLD)


def _calculate_deviation(custom_total: Optional[int] = None,
                         calc: int = 0) -> int:
    return custom_total - calc if custom_total else -calc


def _colorize_deviation_cell(deviation_cell: Cell, deviation: int) -> None:
    if deviation == 0:
        fill_color = RIGHT_COLOR
    else:
        fill_color = WRONG_COLOR
    deviation_cell.fill = PatternFill('solid', fgColor=fill_color)


def _generate_outcome_header(sheet: Worksheet) -> None:
    sheet.insert_rows(1, 2)

    sheet.merge_cells('A1:A2')
    sheet['A1'] = 'Филиал'

    sheet.merge_cells('B1:B2')
    sheet['B1'] = 'Сотрудник'

    sheet.merge_cells('C1:C2')
    sheet['C1'] = 'Налоговая база'

    sheet.merge_cells('D1:E1')
    sheet['D1'] = 'Налог'

    sheet['D2'] = 'Исчислено всего'
    sheet['E2'] = 'Исчислено всего по формуле'

    sheet.merge_cells('F1:F2')
    sheet['F1'] = 'Отклонения'
